import csv
import pandas as pd
from io import StringIO, BytesIO
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
import logging
from decimal import Decimal, InvalidOperation
import re

logger = logging.getLogger(__name__)

class DataExportService:
    def __init__(self):
        self.export_dir = Path("uploads/exports")
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def clean_numeric_value(self, value: Any) -> float:
        """Clean and convert value to numeric, handling various data types"""
        if value is None or value == "" or value == "N/A":
            return 0.0
        
        # If it's already a number
        if isinstance(value, (int, float)):
            return float(value)
        
        # If it's a Decimal
        if isinstance(value, Decimal):
            return float(value)
        
        # If it's a string, clean it
        if isinstance(value, str):
            # Remove common formatting characters but keep digits, dots, and minus
            cleaned = re.sub(r'[^\d.-]', '', value.replace(',', ''))
            if cleaned == '' or cleaned == '-' or cleaned == '.':
                return 0.0
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        
        return 0.0
    
    def debug_column_detection(self, data: List[Dict[str, Any]]) -> Dict[str, bool]:
        """Debug method to see which columns would be detected as numeric"""
        if not data:
            return {}
        
        debug_info = {}
        amount_keywords = [
            'amount', 'salary', 'cost', 'price', 'value', 'total', 'subtotal', 'tax', 
            'allowance', 'bonus', 'deduction', 'gross', 'net', 'overtime',
            'stock', 'quantity', 'capacity', 'weight', 'volume', 'hours', 
            'distance', 'mileage', 'fuel'
        ]
        
        for column_name in data[0].keys():
            has_keyword = any(keyword in column_name.lower() for keyword in amount_keywords)
            is_numeric = self.is_numeric_column(data, column_name)
            debug_info[column_name] = {
                'has_amount_keyword': has_keyword,
                'is_numeric': is_numeric,
                'will_be_summed': has_keyword and is_numeric
            }
        
        return debug_info
    
    def is_numeric_column(self, data: List[Dict[str, Any]], column_name: str) -> bool:
        """Check if a column contains primarily numeric data"""
        if not data:
            return False
        
        numeric_count = 0
        total_count = 0
        has_non_zero = False
        
        for row in data:
            value = row.get(column_name)
            if value is not None and value != "":
                total_count += 1
                try:
                    cleaned_value = self.clean_numeric_value(value)
                    # Check if it's a valid number (including 0)
                    if isinstance(value, (int, float, Decimal)) or (isinstance(value, str) and value.replace('.', '').replace('-', '').replace(',', '').isdigit()):
                        numeric_count += 1
                        if cleaned_value != 0.0:
                            has_non_zero = True
                    elif str(value).strip() in ['0', '0.0', '0.00']:
                        numeric_count += 1
                except:
                    pass
        
        # Consider it numeric if at least 70% of non-empty values are numeric
        # OR if column name suggests it's numeric
        is_percentage_numeric = total_count > 0 and (numeric_count / total_count) >= 0.7
        
        # For debugging
        logger.info(f"Column '{column_name}': {numeric_count}/{total_count} numeric values, has_non_zero: {has_non_zero}")
        
        return is_percentage_numeric
    
    def prepare_data_for_export(self, data: List[Any], fields_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """Prepare data for export by mapping fields and formatting"""
        exported_data = []
        
        for item in data:
            row = {}
            for field_key, display_name in fields_mapping.items():
                try:
                    # Handle different data types
                    if isinstance(item, dict):
                        # For dictionary data (like reports)
                        value = item.get(field_key, None)
                    else:
                        # Handle nested attributes (e.g., 'department.name')
                        if '.' in field_key:
                            value = item
                            for attr in field_key.split('.'):
                                value = getattr(value, attr, None) if value else None
                        else:
                            value = getattr(item, field_key, None)
                    
                    # Format specific data types
                    if value is not None:
                        if isinstance(value, datetime):
                            value = value.strftime("%Y-%m-%d %H:%M:%S")
                        elif isinstance(value, bool):
                            value = "Yes" if value else "No"
                        elif hasattr(value, '__dict__') and not isinstance(value, (str, int, float, Decimal)):  # Complex object
                            value = str(value)
                    else:
                        value = ""
                    
                    row[display_name] = value
                except Exception as e:
                    logger.warning(f"Error accessing field {field_key}: {e}")
                    row[display_name] = ""
            
            exported_data.append(row)
        
        return exported_data
    
    def export_to_csv(self, data: List[Dict[str, Any]], filename: str) -> StreamingResponse:
        """Export data to CSV format"""
        try:
            output = StringIO()
            
            if data:
                fieldnames = list(data[0].keys())
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
            
            output.seek(0)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
            )
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to export data to CSV"
            )
        
    def export_to_excel(self, data: List[Dict[str, Any]], filename: str, sheet_name: str = "Data") -> StreamingResponse:
        """Export data to Excel format with table formatting, borders, and summation"""
        try:
            output = BytesIO()
            
            if data:
                # Clean and prepare numeric data
                cleaned_data = []
                for row in data:
                    cleaned_row = {}
                    for key, value in row.items():
                        cleaned_row[key] = value
                    cleaned_data.append(cleaned_row)
                
                df = pd.DataFrame(cleaned_data)
                
                # Identify amount/numeric columns for summation with more specific keywords
                amount_keywords = [
                    'amount', 'salary', 'cost', 'price', 'value', 'total', 'subtotal', 'tax', 
                    'allowance', 'bonus', 'deduction', 'gross', 'net', 'overtime',
                    'stock', 'quantity', 'capacity', 'weight', 'volume', 'hours', 
                    'distance', 'mileage', 'fuel'
                ]
                
                # Find numeric columns to sum - improved detection
                sum_columns = {}
                sum_column_names = []
                
                for col_idx, column_name in enumerate(df.columns, 1):
                    column_lower = column_name.lower()
                    # Check if column name contains amount-related keywords
                    if any(keyword in column_lower for keyword in amount_keywords):
                        # Verify it's actually numeric
                        if self.is_numeric_column(cleaned_data, column_name):
                            sum_columns[col_idx] = column_name
                            sum_column_names.append(column_name)
                            # Convert column to numeric for proper summation
                            df[column_name] = df[column_name].apply(lambda x: self.clean_numeric_value(x))
                
                logger.info(f"Detected columns for summation: {sum_column_names}")
                
                # Create Excel writer with styling
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get workbook and worksheet for styling
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Import styling classes
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    
                    # Define styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill("solid", fgColor="366092")
                    
                    # Border style
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Sum row styling
                    sum_font = Font(bold=True, size=12)
                    sum_fill = PatternFill("solid", fgColor="D9D9D9")
                    
                    # Get data range
                    max_row = len(df) + 1  # +1 for header
                    max_col = len(df.columns)
                    
                    # Apply header styling
                    for col in range(1, max_col + 1):
                        header_cell = worksheet.cell(row=1, column=col)
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                        header_cell.border = thin_border
                    
                    # Apply borders to all data cells and format numeric columns
                    for row in range(1, max_row + 1):
                        for col in range(1, max_col + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.border = thin_border
                            
                            # Format numeric columns (skip header row)
                            if row > 1 and col in sum_columns:
                                cell.number_format = '#,##0.00'
                    
                    # Add sum row - ALWAYS add it if we have numeric columns
                    if sum_columns:
                        sum_row_number = max_row + 2  # Leave one empty row
                        
                        logger.info(f"Adding sum row at row {sum_row_number}")
                        
                        # Add "TOTAL" label in first column
                        total_label_cell = worksheet.cell(row=sum_row_number, column=1, value="TOTAL")
                        total_label_cell.font = sum_font
                        total_label_cell.fill = sum_fill
                        total_label_cell.border = thin_border
                        
                        # Calculate and add sum values for each numeric column
                        for col_idx, col_name in sum_columns.items():
                            # Calculate sum from the actual data
                            column_sum = 0.0
                            for row_data in cleaned_data:
                                value = row_data.get(col_name, 0)
                                numeric_value = self.clean_numeric_value(value)
                                column_sum += numeric_value
                            
                            # Add sum to the cell
                            sum_cell = worksheet.cell(row=sum_row_number, column=col_idx, value=column_sum)
                            sum_cell.font = sum_font
                            sum_cell.fill = sum_fill
                            sum_cell.number_format = '#,##0.00'
                            sum_cell.border = thin_border
                            
                            logger.info(f"Column '{col_name}' sum: {column_sum}")
                        
                        # Fill empty cells in sum row with borders
                        for col in range(1, max_col + 1):
                            if col not in sum_columns and col != 1:  # Skip total label column
                                empty_cell = worksheet.cell(row=sum_row_number, column=col, value="")
                                empty_cell.fill = sum_fill
                                empty_cell.border = thin_border
                    
                    # Auto-adjust column widths
                    for col_idx, column in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column_letter = get_column_letter(col_idx)
                        
                        for cell in column:
                            try:
                                cell_value = str(cell.value) if cell.value is not None else ""
                                if len(cell_value) > max_length:
                                    max_length = len(cell_value)
                            except:
                                pass
                        
                        # Set minimum width and maximum width
                        adjusted_width = min(max(max_length + 2, 10), 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                logger.info(f"Excel export completed successfully with {len(sum_columns)} numeric columns summed")
            
            output.seek(0)
            
            return StreamingResponse(
                BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
            )
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to export data to Excel"
            )
    
# Export field mappings for different entities
EXPORT_FIELD_MAPPINGS = {
    # Auth & User Management
    "users": {
        "id": "ID",
        "username": "Username", 
        "email": "Email",
        "full_name": "Full Name",
        "phone": "Phone",
        "is_active": "Active",
        "is_verified": "Verified",
        "is_superuser": "Super User",
        "created_at": "Created Date",
        "last_login": "Last Login"
    },
    "roles": {
        "id": "ID",
        "name": "Role Name",
        "description": "Description",
        "is_active": "Active",
        "is_system": "System Role",
        "created_at": "Created Date"
    },
    
    # Organization Management
    "departments": {
        "id": "ID", 
        "name": "Department Name",
        "code": "Department Code",
        "description": "Description",
        "manager_id": "Manager ID",
        "location.name": "Location",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "locations": {
        "id": "ID",
        "name": "Location Name", 
        "code": "Location Code",
        "address": "Address",
        "city": "City",
        "state": "State",
        "country": "Country",
        "postal_code": "Postal Code",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    
    # HR Management
    "employees": {
        "id": "ID",
        "employee_id": "Employee ID",
        "first_name": "First Name", 
        "last_name": "Last Name",
        "email": "Email",
        "phone": "Phone",
        "position": "Position",
        "department.name": "Department",
        "location.name": "Location", 
        "hire_date": "Hire Date",
        "basic_salary": "Basic Salary",
        "is_manager": "Is Manager",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "shifts": {
        "id": "ID",
        "name": "Shift Name",
        "start_time": "Start Time",
        "end_time": "End Time", 
        "break_duration": "Break Duration",
        "late_grace_minutes": "Late Grace Minutes",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "attendance": {
        "id": "ID",
        "employee.employee_id": "Employee ID",
        "employee.first_name": "First Name",
        "employee.last_name": "Last Name",
        "attendance_date": "Date",
        "check_in_time": "Check In",
        "check_out_time": "Check Out", 
        "status": "Status",
        "total_hours": "Total Hours",
        "overtime_hours": "Overtime Hours",
        "late_minutes": "Late Minutes",
        "created_at": "Created Date"
    },
    "deduction_types": {
        "id": "ID",
        "name": "Deduction Type",
        "description": "Description",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "employee_deductions": {
        "id": "ID",
        "employee.employee_id": "Employee ID",
        "employee.first_name": "First Name", 
        "employee.last_name": "Last Name",
        "deduction_type.name": "Deduction Type",
        "total_amount": "Total Amount",
        "paid_amount": "Paid Amount",
        "remaining_amount": "Remaining Amount",
        "monthly_deduction_limit": "Monthly Limit",
        "effective_from": "Effective From",
        "effective_to": "Effective To",
        "status": "Status",
        "description": "Description",
        "created_at": "Created Date"
    },
    "salaries": {
        "id": "ID",
        "employee.employee_id": "Employee ID",
        "employee.first_name": "First Name", 
        "employee.last_name": "Last Name",
        "salary_month": "Salary Month",
        "basic_salary": "Basic Salary",
        "housing_allowance": "Housing Allowance",
        "transport_allowance": "Transport Allowance",
        "overtime_amount": "Overtime Amount",
        "gross_salary": "Gross Salary",
        "total_deductions": "Total Deductions",
        "net_salary": "Net Salary",
        "payment_status": "Payment Status",
        "created_at": "Created Date"
    },
    "holidays": {
        "id": "ID", 
        "name": "Holiday Name",
        "date": "Date",
        "description": "Description",
        "is_recurring": "Is Recurring",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    
    # Inventory Management
    "items": {
        "id": "ID",
        "item_code": "Item Code",
        "name": "Item Name",
        "description": "Description", 
        "category.name": "Category",
        "stock_type.name": "Stock Type",
        "unit_type": "Unit Type",
        "unit_cost": "Unit Cost",
        "selling_price": "Selling Price",
        "minimum_stock_level": "Min Stock Level",
        "maximum_stock_level": "Max Stock Level", 
        "reorder_point": "Reorder Point",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "categories": {
        "id": "ID",
        "name": "Category Name",
        "description": "Description",
        "parent.name": "Parent Category",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "stock_types": {
        "id": "ID",
        "name": "Stock Type Name",
        "description": "Description",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "stock_levels": {
        "id": "ID",
        "item.item_code": "Item Code",
        "item.name": "Item Name",
        "location.name": "Location",
        "current_stock": "Current Stock",
        "reserved_stock": "Reserved Stock",
        "available_stock": "Available Stock",
        "par_level_min": "Min Level",
        "par_level_max": "Max Level",
        "updated_at": "Last Updated"
    },
    "stock_movements": {
        "id": "ID",
        "item.item_code": "Item Code",
        "item.name": "Item Name",
        "location.name": "Location",
        "movement_type": "Movement Type",
        "quantity": "Quantity",
        "unit_cost": "Unit Cost",
        "reference_type": "Reference Type",
        "reference_id": "Reference ID",
        "batch_number": "Batch Number",
        "movement_date": "Movement Date",
        "remarks": "Remarks"
    },
    "reorder_requests": {
        "id": "ID",
        "request_number": "Request Number",
        "location.name": "Location", 
        "status": "Status",
        "priority": "Priority",
        "request_date": "Request Date",
        "required_date": "Required Date",
        "total_estimated_cost": "Estimated Cost",
        "created_at": "Created Date"
    },
    "transfers": {
        "id": "ID",
        "transfer_number": "Transfer Number",
        "from_location.name": "From Location",
        "to_location.name": "To Location",
        "status": "Status",
        "transfer_date": "Transfer Date",
        "expected_date": "Expected Date",
        "created_at": "Created Date"
    },
    "inventory_counts": {
        "id": "ID",
        "count_number": "Count Number",
        "location.name": "Location",
        "count_date": "Count Date",
        "count_type": "Count Type",
        "status": "Status", 
        "notes": "Notes",
        "created_at": "Created Date"
    },
    
    # Purchase Management
    "purchase_orders": {
        "id": "ID",
        "po_number": "PO Number",
        "supplier.name": "Supplier",
        "status": "Status",
        "order_date": "Order Date",
        "expected_delivery_date": "Expected Delivery",
        "subtotal": "Subtotal",
        "tax_amount": "Tax Amount",
        "total_amount": "Total Amount",
        "approved_date": "Approved Date",
        "created_at": "Created Date"
    },
    "suppliers": {
        "id": "ID",
        "name": "Supplier Name",
        "supplier_code": "Supplier Code",
        "contact_person": "Contact Person",
        "email": "Email",
        "phone": "Phone", 
        "address": "Address",
        "city": "City",
        "country": "Country",
        "is_active": "Active",
        "created_at": "Created Date"
    },
    "goods_receipts": {
        "id": "ID",
        "receipt_number": "Receipt Number",
        "purchase_order.po_number": "PO Number",
        "supplier.name": "Supplier",
        "receipt_date": "Receipt Date",
        "delivered_by": "Delivered By",
        "notes": "Notes",
        "created_at": "Created Date"
    },
    
    # Logistics Management  
    "shipments": {
        "id": "ID",
        "shipment_number": "Shipment Number",
        "from_location.name": "From Location",
        "to_location.name": "To Location",
        "status": "Status",
        "driver.employee.first_name": "Driver First Name",
        "driver.employee.last_name": "Driver Last Name",
        "vehicle.vehicle_number": "Vehicle Number",
        "shipment_date": "Shipment Date",
        "expected_delivery_date": "Expected Delivery",
        "distance_km": "Distance (KM)",
        "created_at": "Created Date"
    },
    "drivers": {
        "id": "ID",
        "employee.first_name": "First Name",
        "employee.last_name": "Last Name",
        "employee.employee_id": "Employee ID",
        "license_number": "License Number",
        "license_type": "License Type",
        "license_expiry": "License Expiry",
        "phone": "Phone",
        "is_active": "Active",
        "is_available": "Available",
        "created_at": "Created Date"
    },
    "vehicles": {
        "id": "ID", 
        "vehicle_number": "Vehicle Number",
        "vehicle_type": "Vehicle Type",
        "make": "Make",
        "model": "Model",
        "year": "Year",
        "capacity_weight": "Weight Capacity",
        "capacity_volume": "Volume Capacity",
        "fuel_type": "Fuel Type",
        "current_mileage": "Current Mileage",
        "is_active": "Active",
        "is_available": "Available",
        "created_at": "Created Date"
    },

    # Reports
    "stock_levels_report": {
        "sl": "SL",
        "sku": "SKU",
        "item_name": "Item Name",
        "category": "Category",
        "location": "Location",
        "current_stock": "Current Stock",
        "available_stock": "Available Stock",
        "reserved_stock": "Reserved Stock",
        "minimum_stock": "Min Stock",
        "maximum_stock": "Max Stock",
        "reorder_point": "Reorder Point",
        "unit_cost": "Unit Cost",
        "stock_value": "Stock Value",
        "stock_status": "Stock Status",
        "priority": "Priority",
        "unit": "Unit"
    },
    "stock_movements_report": {
        "sl": "SL",
        "sku": "SKU",
        "item_name": "Item Name",
        "location": "Location",
        "movement_type": "Movement Type",
        "quantity": "Quantity",
        "unit_cost": "Unit Cost",
        "total_value": "Total Value",
        "reference_type": "Reference Type",
        "reference_id": "Reference ID",
        "batch_number": "Batch Number",
        "expiry_date": "Expiry Date",
        "remarks": "Remarks",
        "movement_date": "Movement Date",
        "transaction_direction": "Direction"
    },
    "low_stock_alerts_report": {
        "sl": "SL",
        "sku": "SKU",
        "item_name": "Item Name",
        "category": "Category",
        "location": "Location",
        "current_stock": "Current Stock",
        "reorder_point": "Reorder Point",
        "minimum_stock": "Min Stock",
        "shortage_quantity": "Shortage Qty",
        "priority": "Priority",
        "unit": "Unit",
        "unit_cost": "Unit Cost",
        "estimated_reorder_cost": "Estimated Cost",
        "days_out_of_stock": "Days Until Out"
    },
    "purchase_orders_summary_report": {
        "sl": "SL",
        "po_number": "PO Number",
        "supplier_name": "Supplier",
        "order_date": "Order Date",
        "expected_delivery_date": "Expected Delivery",
        "status": "Status",
        "total_amount": "Total Amount",
        "total_items": "Total Items",
        "total_quantity": "Total Quantity",
        "total_received": "Total Received",
        "receipt_status": "Receipt Status",
        "completion_percentage": "Completion %",
        "approved_date": "Approved Date",
        "days_pending": "Days Pending"
    },
    "demand_forecast_report": {
        "sl": "SL",
        "sku": "SKU",
        "productName": "Product Name",
        "category": "Category",
        "location": "Location",
        "forecastPeriod": "Forecast Period",
        "currentSalesData": "Current Sales",
        "plannedSalesTarget": "Sales Target",
        "demandVariationPercentage": "Demand Variation %",
        "averageDailyDemand": "Avg Daily Demand",
        "transactionCount": "Transaction Count",
        "forecastAccuracy": "Forecast Accuracy %"
    },
    "attendance_summary_report": {
        "sl": "SL",
        "employee_id": "Employee ID",
        "employee_code": "Employee Code",
        "employee_name": "Employee Name",
        "department": "Department",
        "location": "Location",
        "total_working_days": "Working Days",
        "present_days": "Present Days",
        "absent_days": "Absent Days",
        "late_days": "Late Days",
        "early_leave_days": "Early Leave Days",
        "attendance_percentage": "Attendance %",
        "punctuality_percentage": "Punctuality %",
        "total_hours_worked": "Total Hours",
        "overtime_hours": "Overtime Hours",
        "average_daily_hours": "Avg Daily Hours",
        "total_late_minutes": "Late Minutes",
        "total_early_leave_minutes": "Early Leave Minutes",
        "overall_status": "Status",
        "productivity_score": "Productivity Score"
    },
    "salary_summary_report": {
        "sl": "SL",
        "employee_id": "Employee Code",
        "employee_name": "Employee Name",
        "department": "Department",
        "location": "Location",
        "designation": "Designation",
        "salary_month": "Salary Month",
        "basic_salary": "Basic Salary",
        "housing_allowance": "Housing Allowance",
        "transport_allowance": "Transport Allowance",
        "overtime_amount": "Overtime Amount",
        "bonus": "Bonus",
        "gross_salary": "Gross Salary",
        "total_deductions": "Total Deductions",
        "late_deductions": "Late Deductions",
        "absent_deductions": "Absent Deductions",
        "other_deductions": "Other Deductions",
        "deduction_percentage": "Deduction %",
        "net_salary": "Net Salary",
        "working_days": "Working Days",
        "present_days": "Present Days",
        "absent_days": "Absent Days",
        "late_days": "Late Days",
        "attendance_efficiency": "Attendance Efficiency %",
        "payment_status": "Payment Status",
        "salary_status": "Status",
        "payment_date": "Payment Date",
        "payment_method": "Payment Method",
        "generated_date": "Generated Date",
        "salary_efficiency_score": "Efficiency Score",
        "cost_per_day": "Cost Per Day"
    },
    "shipment_tracking_report": {
        "sl": "SL",
        "shipment_number": "Shipment Number",
        "from_location": "From Location",
        "to_location": "To Location",
        "driver_name": "Driver Name",
        "driver_phone": "Driver Phone",
        "vehicle_number": "Vehicle Number",
        "vehicle_type": "Vehicle Type",
        "shipment_date": "Shipment Date",
        "expected_delivery": "Expected Delivery",
        "actual_delivery": "Actual Delivery",
        "pickup_time": "Pickup Time",
        "delivery_time": "Delivery Time",
        "status": "Status",
        "delivery_performance": "Delivery Performance",
        "delay_days": "Delay Days",
        "urgency": "Urgency",
        "transit_time_hours": "Transit Time (Hours)",
        "total_items": "Total Items",
        "total_quantity": "Total Quantity",
        "total_delivered": "Total Delivered",
        "completion_percentage": "Completion %",
        "distance_km": "Distance (KM)",
        "total_weight": "Total Weight",
        "total_volume": "Total Volume",
        "fuel_cost": "Fuel Cost",
        "pickup_verified": "Pickup Verified",
        "delivery_verified": "Delivery Verified",
        "reference_type": "Reference Type",
        "reference_id": "Reference ID",
        "notes": "Notes",
        "cost_per_km": "Cost Per KM",
        "efficiency_score": "Efficiency Score"
    }
}